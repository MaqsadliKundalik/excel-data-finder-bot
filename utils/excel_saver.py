from openpyxl import load_workbook
from database.models import Medicines
from decimal import Decimal, InvalidOperation

async def load_excel_to_database(file_path: str, progress_callback=None):
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        
        updated_count = 0
        created_count = 0
        read_rows = 0
        skipped_rows = 0
        
        total_rows = sheet.max_row - 1
        seen_numbers = {}
        
        for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
            if not any(row):
                continue
            
            read_rows += 1
            
            if progress_callback and total_rows > 0:
                percent = int((index / total_rows) * 100)
                await progress_callback(percent, index, total_rows)
                
            trade_name = str(row[0]) if len(row) > 0 and row[0] else ""
            mnn = str(row[1]) if len(row) > 1 and row[1] else ""
            manufacturer = str(row[2]) if len(row) > 2 and row[2] else ""
            form = str(row[3]) if len(row) > 3 and row[3] else ""
            registration_number = str(row[4]) if len(row) > 4 and row[4] else ""
            
            if registration_number in seen_numbers:
                skipped_rows += 1
                continue
            
            seen_numbers[registration_number] = True
            
            try:
                if len(row) > 5 and row[5]:
                    price_value = str(row[5]).replace(',', '.').strip()
                    price = Decimal(price_value) if price_value else Decimal('0')
                else:
                    price = Decimal('0')
            except (InvalidOperation, ValueError):
                price = Decimal('0')
            
            dispensing_mode = str(row[6]) if len(row) > 6 and row[6] else ""
            
            existing_medicine = await Medicines.get_or_none(
                registration_number=registration_number
            )
            
            if existing_medicine:
                existing_medicine.trade_name = trade_name
                existing_medicine.mnn = mnn
                existing_medicine.manufacturer = manufacturer
                existing_medicine.form = form
                existing_medicine.price = price
                existing_medicine.dispensing_mode = dispensing_mode
                await existing_medicine.save()
                updated_count += 1
            else:
                await Medicines.create(
                    trade_name=trade_name,
                    mnn=mnn,
                    manufacturer=manufacturer,
                    form=form,
                    registration_number=registration_number,
                    price=price,
                    dispensing_mode=dispensing_mode
                )
                created_count += 1
        
        return {
            'success': True,
            'message': "✅ Маълумотлар муваффақиятли сақланди!"
        }
        
    except FileNotFoundError:
        return {
            'success': False,
            'message': f"Файл топилмади: {file_path}"
        }
    except Exception as e:
        return {
            'success': False,
            'message': f"Хатолик юз берди: {str(e)}"
        }

